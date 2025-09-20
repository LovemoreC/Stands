from fastapi import FastAPI, HTTPException, Depends, Header, Request, UploadFile, File, Form, Response
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import Any, Dict, List, Optional
from datetime import datetime, timedelta
import hashlib
import hmac
import csv
import base64
import uuid
import os
import logging
from io import BytesIO
import re
import jwt
from passlib.context import CryptContext
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
from starlette.datastructures import FormData, UploadFile as StarletteUploadFile
from .database import init_db, get_session, SessionLocal
from .mailer import (
    MailRequest,
    build_attachments,
    get_mailer,
    EmailSendError,
)
from .repositories import Repositories
from .projects import ProjectsService
from .reporting import (
    generate_properties_report,
    generate_mandates_report,
    generate_loans_report,
    stream_csv,
)
from .models import (
    Project,
    ProjectCreate,
    Stand,
    StandCreate,
    PropertyStatus,
    Mandate,
    MandateRecord,
    MandateHistoryEntry,
    Agent,
    AgentCreate,
    AgentInDB,
    AgentRole,
    MandateStatus,
    SubmissionStatus,
    Offer,
    PropertyApplication,
    AccountOpening,
    StatusUpdate,
    AccountSetup,
    Deposit,
    LoanApplication,
    LoanDecisionUpdate,
    LoanDecision,
    Loan,
    LoanDecisionRequest,
    LoanStatus,
    Agreement,
    AgreementCreate,
    AgreementUpload,
    AgreementExecution,
    AgreementSign,
    AgreementStatus,
    UploadedFile,
    DocumentRequirement,
    DocumentWorkflow,
    ImportedDepositAccount,
    ImportedLoanAccount,
    ContactSettings,
    ContactSettingsResponse,
    CustomerProfile,
    LoanTeamReply,
)

logger = logging.getLogger(__name__)

DEFAULT_FRONTEND_ORIGINS = ["http://localhost:5173"]


def resolve_frontend_origins() -> List[str]:
    raw_origins = os.environ.get("FRONTEND_ORIGINS")
    if not raw_origins:
        return DEFAULT_FRONTEND_ORIGINS.copy()
    parsed_origins = [origin.strip() for origin in raw_origins.split(",") if origin.strip()]
    return parsed_origins or DEFAULT_FRONTEND_ORIGINS.copy()


SECRET_KEY = os.environ["SECRET_KEY"]
INITIAL_ADMIN_TOKEN = os.environ["INITIAL_ADMIN_TOKEN"]

FRONTEND_ORIGINS = resolve_frontend_origins()


def _parse_recipients(value: Optional[str]) -> List[str]:
    if not value:
        return []
    return [part.strip() for part in value.split(",") if part.strip()]


DEFAULT_DEPOSIT_ACCOUNTS_RECIPIENTS = _parse_recipients(
    os.environ.get("DEPOSIT_ACCOUNTS_EMAIL")
)
DEPOSIT_CONTACT_SETTINGS_KEY = "deposit_notifications"
MAIL_MAX_RETRIES = max(1, int(os.environ.get("MAIL_MAX_RETRIES", "3")))

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

app = FastAPI(title="Property Management API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=FRONTEND_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def seed_initial_admin_if_configured() -> None:
    username = os.environ.get("INITIAL_ADMIN_USERNAME")
    if not username:
        return

    password = os.environ.get("INITIAL_ADMIN_PASSWORD")
    password_hash = os.environ.get("INITIAL_ADMIN_PASSWORD_HASH")

    if not password and not password_hash:
        logger.warning(
            "INITIAL_ADMIN_USERNAME is set but neither INITIAL_ADMIN_PASSWORD nor "
            "INITIAL_ADMIN_PASSWORD_HASH was provided; skipping admin seeding"
        )
        return

    session = SessionLocal()
    try:
        repos = Repositories(session)
        if repos.agents.list():
            logger.debug(
                "Skipping initial admin seeding because at least one agent already exists"
            )
            return

        hashed = password_hash
        if password:
            hashed = hash_password(password)

        if not hashed:
            logger.warning(
                "Unable to determine password hash for initial admin; skipping seeding"
            )
            return

        admin = AgentInDB(
            username=username,
            role=AgentRole.ADMIN,
            password_hash=hashed,
        )
        repos.agents.add(admin)
        logger.info("Seeded initial admin '%s' from environment", username)
    finally:
        session.close()


def run_startup_tasks() -> None:
    logger.info("Configured frontend origins: %s", ", ".join(FRONTEND_ORIGINS))
    init_db()
    seed_initial_admin_if_configured()


@app.on_event("startup")
def on_startup() -> None:
    run_startup_tasks()


def get_repositories(session=Depends(get_session)) -> Repositories:
    return Repositories(session)


def get_projects_service(
    repos: Repositories = Depends(get_repositories),
) -> ProjectsService:
    return ProjectsService(repos)


def _normalize_recipients(recipients: List[str]) -> List[str]:
    normalized: List[str] = []
    seen: set[str] = set()
    for recipient in recipients:
        trimmed = str(recipient).strip()
        if not trimmed:
            continue
        key = trimmed.lower()
        if key in seen:
            continue
        normalized.append(trimmed)
        seen.add(key)
    return normalized


def _parse_loan_team_decision(body: str) -> LoanDecision | None:
    normalized = body.lower()
    if "approved" in normalized:
        return LoanDecision.APPROVED
    if "declined" in normalized or "rejected" in normalized:
        return LoanDecision.REJECTED
    return None


def _extract_recipients(data: Any) -> List[str]:
    if not isinstance(data, dict):
        return []
    raw = data.get("recipients")
    if not isinstance(raw, list):
        return []
    return _normalize_recipients([str(value) for value in raw])


def _resolve_deposit_recipients(repos: Repositories) -> List[str]:
    stored = repos.contact_settings.get(DEPOSIT_CONTACT_SETTINGS_KEY)
    if stored is not None:
        return _extract_recipients(stored)
    return _normalize_recipients(DEFAULT_DEPOSIT_ACCOUNTS_RECIPIENTS)


def _get_contact_settings_response(repos: Repositories) -> ContactSettingsResponse:
    stored = repos.contact_settings.get(DEPOSIT_CONTACT_SETTINGS_KEY)
    if stored is None:
        return ContactSettingsResponse(
            recipients=_normalize_recipients(DEFAULT_DEPOSIT_ACCOUNTS_RECIPIENTS),
            configured=False,
        )
    return ContactSettingsResponse(recipients=_extract_recipients(stored))


def _store_contact_settings(
    repos: Repositories, payload: ContactSettings
) -> ContactSettingsResponse:
    normalized = _normalize_recipients([str(item) for item in payload.recipients])
    repos.contact_settings.set(
        DEPOSIT_CONTACT_SETTINGS_KEY,
        {"recipients": normalized},
    )
    return ContactSettingsResponse(recipients=normalized)


def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(password: str, hashed: str) -> bool:
    if hashed.startswith("$2"):
        return pwd_context.verify(password, hashed)
    legacy = hashlib.sha256(password.encode()).hexdigest()
    return hmac.compare_digest(legacy, hashed)


def _send_submission_email(
    subject: str,
    metadata: Dict[str, Optional[str | int]],
    primary_document: Optional[UploadedFile],
    required_documents: Dict[str, UploadedFile] | None,
    repos: Repositories,
) -> None:
    recipients = _resolve_deposit_recipients(repos)
    if not recipients:
        logger.debug(
            "Skipping email dispatch for '%s' because no recipients were configured",
            subject,
        )
        return

    attachments = []
    if primary_document and getattr(primary_document, "content", None):
        try:
            attachments.extend(build_attachments([primary_document]))
        except EmailSendError as exc:
            logger.warning("Unable to attach primary document for '%s': %s", subject, exc)
    if required_documents:
        try:
            attachments.extend(build_attachments(required_documents.values()))
        except EmailSendError as exc:
            logger.warning(
                "Unable to attach required documents for '%s': %s",
                subject,
                exc,
            )

    lines = [subject, ""]
    for key, value in metadata.items():
        normalized = value if value not in (None, "") else "N/A"
        lines.append(f"{key}: {normalized}")

    request = MailRequest(
        to=recipients,
        subject=subject,
        body="\n".join(lines),
        attachments=attachments,
    )

    mailer = get_mailer()
    last_error: Exception | None = None
    for attempt in range(1, MAIL_MAX_RETRIES + 1):
        try:
            mailer.send(request)
            return
        except EmailSendError as exc:
            last_error = exc
            logger.exception(
                "Email dispatch failed for '%s' on attempt %d/%d",
                subject,
                attempt,
                MAIL_MAX_RETRIES,
            )
        except Exception as exc:  # pragma: no cover - defensive
            last_error = exc
            logger.exception(
                "Unexpected error while sending '%s' on attempt %d/%d",
                subject,
                attempt,
                MAIL_MAX_RETRIES,
            )
    if last_error:
        message = f"Email delivery failed for {subject}: {last_error}"
        repos.notifications.append(message)


def create_access_token(data: dict, expires_delta: timedelta | None = None) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(hours=1))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm="HS256")


def get_current_agent(
    authorization: str = Header(...), repos: Repositories = Depends(get_repositories)
) -> Agent:
    if not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Invalid authentication token")
    token = authorization.split(" ", 1)[1]
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication token")
    username = payload.get("sub")
    role = payload.get("role")
    if not username or not role:
        raise HTTPException(status_code=401, detail="Invalid authentication token")
    agent_in_db = repos.agents.get(username)
    if not agent_in_db:
        raise HTTPException(status_code=401, detail="Invalid authentication token")
    try:
        token_role = AgentRole(role)
    except ValueError:
        raise HTTPException(status_code=401, detail="Invalid authentication token")
    if agent_in_db.role != token_role:
        raise HTTPException(status_code=401, detail="Invalid authentication token")
    return Agent(username=agent_in_db.username, role=agent_in_db.role)


def require_admin(agent: Agent = Depends(get_current_agent)) -> Agent:
    if agent.role != AgentRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin privileges required")
    return agent


def require_management(agent: Agent = Depends(get_current_agent)) -> Agent:
    if agent.role not in (AgentRole.ADMIN, AgentRole.MANAGER):
        raise HTTPException(status_code=403, detail="Management privileges required")
    return agent


def require_compliance(agent: Agent = Depends(get_current_agent)) -> Agent:
    if agent.role not in (AgentRole.COMPLIANCE, AgentRole.ADMIN):
        raise HTTPException(status_code=403, detail="Compliance privileges required")
    return agent


@app.middleware("http")
async def log_request(request: Request, call_next):
    auth = request.headers.get("authorization")
    username = "anonymous"
    if auth and auth.lower().startswith("bearer "):
        token = auth.split(" ", 1)[1]
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            username = payload.get("sub", "anonymous")
        except jwt.PyJWTError:
            pass
    session = SessionLocal()
    repos = Repositories(session)
    timestamp = datetime.utcnow().isoformat()
    status_code = 500
    response: Response | None = None
    try:
        response = await call_next(request)
        status_code = response.status_code
    except Exception:
        raise
    finally:
        repos.audit_log.append(
            f"{timestamp} - {username} - {request.method} {request.url.path} - {status_code}"
        )
        session.close()
    return response


@app.post("/agents", response_model=Agent)
def create_agent(
    agent: AgentCreate,
    authorization: str = Header(None),
    bootstrap_token: str | None = Header(None, alias="X-Bootstrap-Token"),
    repos: Repositories = Depends(get_repositories),
):
    existing_agents = repos.agents.list()
    if existing_agents:
        if not authorization:
            raise HTTPException(status_code=401, detail="Missing authentication token")
        current = get_current_agent(authorization=authorization, repos=repos)
        if current.role != AgentRole.ADMIN:
            raise HTTPException(status_code=403, detail="Admin privileges required")
    else:
        if not bootstrap_token:
            raise HTTPException(status_code=401, detail="Missing bootstrap token")
        if not hmac.compare_digest(bootstrap_token, INITIAL_ADMIN_TOKEN):
            raise HTTPException(status_code=403, detail="Invalid bootstrap token")
    if repos.agents.get(agent.username):
        raise HTTPException(status_code=400, detail="Agent exists")
    if agent.role not in AgentRole:
        raise HTTPException(status_code=400, detail="Unknown role")
    if len(agent.password) < 8 or not agent.password.strip():
        raise HTTPException(status_code=400, detail="Password does not meet requirements")
    password_hash = hash_password(agent.password)
    agent_db = AgentInDB(username=agent.username, role=agent.role, password_hash=password_hash)
    repos.agents.add(agent_db)
    return Agent(username=agent.username, role=agent.role)


@app.get("/agents", response_model=List[Agent])
def list_agents(
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    return [Agent(username=agent.username, role=agent.role) for agent in repos.agents.list()]


@app.get("/contact-settings/deposit", response_model=ContactSettingsResponse)
def get_deposit_contact_settings(
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    return _get_contact_settings_response(repos)


@app.post("/contact-settings/deposit", response_model=ContactSettingsResponse)
def create_deposit_contact_settings(
    payload: ContactSettings,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    if repos.contact_settings.get(DEPOSIT_CONTACT_SETTINGS_KEY) is not None:
        raise HTTPException(status_code=409, detail="Contact settings already configured")
    return _store_contact_settings(repos, payload)


@app.put("/contact-settings/deposit", response_model=ContactSettingsResponse)
def update_deposit_contact_settings(
    payload: ContactSettings,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    return _store_contact_settings(repos, payload)


@app.delete("/contact-settings/deposit", status_code=204)
def delete_deposit_contact_settings(
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    repos.contact_settings.delete(DEPOSIT_CONTACT_SETTINGS_KEY)
    return Response(status_code=204)


class LoginRequest(BaseModel):
    username: str
    password: str


class ImportResult(BaseModel):
    imported: int
    errors: List[str]


class AuditEntry(BaseModel):
    timestamp: str
    user: str
    action: str
    status: int


class DocumentRequirementCreate(BaseModel):
    name: str
    applies_to: DocumentWorkflow


class DocumentRequirementUpdate(BaseModel):
    name: Optional[str] = None
    applies_to: Optional[DocumentWorkflow] = None


class DocumentRequirementReorder(BaseModel):
    applies_to: DocumentWorkflow
    ordered_ids: List[int]

@app.get("/document-requirements", response_model=List[DocumentRequirement])
def list_document_requirements(
    applies_to: DocumentWorkflow | None = None,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    if applies_to:
        return _requirements_for_workflow(repos, applies_to)
    requirements = repos.document_requirements.list()
    return sorted(requirements, key=lambda r: (r.applies_to.value, r.order, r.id))


@app.post("/document-requirements", response_model=DocumentRequirement)
def create_document_requirement(
    payload: DocumentRequirementCreate,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    requirement_id = _allocate_requirement_id(repos)
    order = _next_requirement_order(repos, payload.applies_to)
    requirement = DocumentRequirement(
        id=requirement_id,
        name=payload.name,
        applies_to=payload.applies_to,
        slug=_derive_unique_slug(repos, payload.applies_to, payload.name),
        order=order,
    )
    repos.document_requirements.add(requirement)
    return requirement


@app.put("/document-requirements/{requirement_id}", response_model=DocumentRequirement)
def update_document_requirement(
    requirement_id: int,
    payload: DocumentRequirementUpdate,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    requirement = repos.document_requirements.get(requirement_id)
    if not requirement:
        raise HTTPException(status_code=404, detail="Document requirement not found")
    update_data = requirement.model_dump()
    target_workflow = requirement.applies_to
    slug = requirement.slug
    if payload.name is not None:
        update_data["name"] = payload.name
    if payload.applies_to is not None and payload.applies_to != requirement.applies_to:
        target_workflow = payload.applies_to
        update_data["applies_to"] = payload.applies_to
        update_data["order"] = _next_requirement_order(repos, payload.applies_to)
    if target_workflow != requirement.applies_to:
        conflict = any(
            req.slug == slug and req.id != requirement.id
            for req in repos.document_requirements.list()
            if req.applies_to == target_workflow
        )
        if conflict:
            slug = _derive_unique_slug(
                repos,
                target_workflow,
                update_data["name"],
                exclude_id=requirement.id,
            )
    update_data["slug"] = slug
    updated = DocumentRequirement(**update_data)
    repos.document_requirements.add(updated)
    return updated


@app.post("/document-requirements/reorder", response_model=List[DocumentRequirement])
def reorder_document_requirements(
    payload: DocumentRequirementReorder,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    requirements = _requirements_for_workflow(repos, payload.applies_to)
    expected_ids = {req.id for req in requirements}
    if set(payload.ordered_ids) != expected_ids:
        raise HTTPException(status_code=400, detail="Ordered ids must match requirements")
    for index, requirement_id in enumerate(payload.ordered_ids, start=1):
        requirement = repos.document_requirements.get(requirement_id)
        if not requirement or requirement.applies_to != payload.applies_to:
            raise HTTPException(status_code=400, detail="Invalid requirement id")
        requirement.order = index
        repos.document_requirements.add(requirement)
    return _requirements_for_workflow(repos, payload.applies_to)


@app.delete("/document-requirements/{requirement_id}", response_model=DocumentRequirement)
def delete_document_requirement(
    requirement_id: int,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    requirement = repos.document_requirements.get(requirement_id)
    if not requirement:
        raise HTTPException(status_code=404, detail="Document requirement not found")
    repos.document_requirements.delete(requirement_id)
    return requirement


@app.post("/auth/login")
def login(data: LoginRequest, repos: Repositories = Depends(get_repositories)):
    agent = repos.agents.get(data.username)
    if not agent or not verify_password(data.password, agent.password_hash):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if not agent.password_hash.startswith("$2"):
        agent.password_hash = hash_password(data.password)
        repos.agents.add(agent)
    token = create_access_token({"sub": agent.username, "role": agent.role.value})
    return {"token": token, "role": agent.role, "username": agent.username}


@app.post("/projects", response_model=Project)
def create_project(
    project: ProjectCreate,
    _: Agent = Depends(require_admin),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.create_project(project)


@app.get("/projects", response_model=List[Project])
def list_projects(
    _: Agent = Depends(get_current_agent),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.list_projects()


@app.put("/projects/{project_id}", response_model=Project)
def update_project(
    project_id: int,
    project: Project,
    _: Agent = Depends(require_admin),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.update_project(project_id, project)


@app.delete("/projects/{project_id}", response_model=Project)
def delete_project(
    project_id: int,
    _: Agent = Depends(require_admin),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.delete_project(project_id)


@app.get("/projects/{project_id}/stands", response_model=List[Stand])
def list_project_stands(
    project_id: int,
    _: Agent = Depends(get_current_agent),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.list_stands(project_id)


@app.post("/projects/{project_id}/stands", response_model=Stand)
def create_project_stand(
    project_id: int,
    stand: StandCreate,
    _: Agent = Depends(require_admin),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.create_stand(project_id, stand)


@app.put("/projects/{project_id}/stands/{stand_id}", response_model=Stand)
def update_project_stand(
    project_id: int,
    stand_id: int,
    stand: Stand,
    _: Agent = Depends(require_admin),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.update_stand(project_id, stand_id, stand)


@app.delete("/projects/{project_id}/stands/{stand_id}", response_model=Stand)
def delete_project_stand(
    project_id: int,
    stand_id: int,
    _: Agent = Depends(require_admin),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.delete_stand(project_id, stand_id)


@app.post("/stands", response_model=Stand)
def create_stand(
    stand: StandCreate,
    _: Agent = Depends(require_admin),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.create_stand(stand.project_id, stand)


@app.put("/stands/{stand_id}", response_model=Stand)
def update_stand(
    stand_id: int,
    stand: Stand,
    _: Agent = Depends(require_admin),
    service: ProjectsService = Depends(get_projects_service),
):
    return service.update_stand(stand.project_id, stand_id, stand)


@app.delete("/stands/{stand_id}", response_model=Stand)
def archive_stand(
    stand_id: int,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    stand = repos.stands.get(stand_id)
    if not stand:
        raise HTTPException(status_code=404, detail="Stand not found")
    if stand.status == PropertyStatus.SOLD:
        raise HTTPException(status_code=400, detail="Stand already sold")
    stand.status = PropertyStatus.ARCHIVED
    repos.stands.add(stand)
    return stand


@app.post("/stands/{stand_id}/mandate", response_model=Stand)
def assign_mandate(
    stand_id: int,
    mandate: Mandate,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    stand = repos.stands.get(stand_id)
    if not stand:
        raise HTTPException(status_code=404, detail="Stand not found")
    if stand.status == PropertyStatus.SOLD:
        raise HTTPException(status_code=400, detail="Stand already sold")
    stand.mandate = mandate
    stand.mandate.status = MandateStatus.PENDING
    repos.stands.add(stand)
    return stand


@app.put("/stands/{stand_id}/mandate/accept", response_model=Stand)
def accept_mandate(
    stand_id: int,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    stand = repos.stands.get(stand_id)
    if not stand:
        raise HTTPException(status_code=404, detail="Stand not found")
    if stand.status == PropertyStatus.SOLD:
        raise HTTPException(status_code=400, detail="Stand already sold")
    if not stand.mandate or stand.mandate.agent != agent.username:
        raise HTTPException(status_code=403, detail="Not authorized to accept this mandate")
    stand.mandate.status = MandateStatus.ACCEPTED
    repos.stands.add(stand)
    return stand


@app.get("/stands/available", response_model=List[Stand])
def available_stands(
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    result = [s for s in repos.stands.list() if s.status == PropertyStatus.AVAILABLE]
    if agent.role == AgentRole.AGENT:
        result = [
            s
            for s in result
            if s.mandate and s.mandate.agent == agent.username and s.mandate.status == MandateStatus.ACCEPTED
        ]
    return result


@app.get("/stands/{stand_id}", response_model=Stand)
def get_stand(
    stand_id: int,
    _: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    stand = repos.stands.get(stand_id)
    if not stand:
        raise HTTPException(status_code=404, detail="Stand not found")
    return stand


# Mandate endpoints


@app.post("/mandates", response_model=MandateRecord)
def create_mandate(
    mandate: MandateRecord,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    if repos.mandates.get(mandate.id):
        raise HTTPException(status_code=400, detail="Mandate exists")
    if not repos.projects.get(mandate.project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    if not repos.agents.get(mandate.agent):
        raise HTTPException(status_code=404, detail="Agent not found")
    repos.mandates.add(mandate)
    repos.mandate_history.set(
        mandate.id,
        [
            {
                "timestamp": datetime.utcnow().isoformat(),
                "status": mandate.status.value,
            }
        ],
    )
    return mandate


@app.put("/mandates/{mandate_id}", response_model=MandateRecord)
def update_mandate(
    mandate_id: int,
    mandate: MandateRecord,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    existing = repos.mandates.get(mandate_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Mandate not found")
    if mandate.id != mandate_id:
        raise HTTPException(status_code=400, detail="Mandate ID mismatch")
    if not repos.projects.get(mandate.project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    if not repos.agents.get(mandate.agent):
        raise HTTPException(status_code=404, detail="Agent not found")
    if mandate.status != existing.status:
        history = repos.mandate_history.get(mandate_id, [])
        history.append(
            {
                "timestamp": datetime.utcnow().isoformat(),
                "status": mandate.status.value,
            }
        )
        repos.mandate_history.set(mandate_id, history)
    repos.mandates.add(mandate)
    return mandate


@app.get("/mandates", response_model=List[MandateRecord])
def list_mandates(
    _: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    return repos.mandates.list()


@app.get("/mandates/{mandate_id}/history", response_model=List[MandateHistoryEntry])
def get_mandate_history(
    mandate_id: int,
    _: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if not repos.mandates.get(mandate_id):
        raise HTTPException(status_code=404, detail="Mandate not found")
    history = repos.mandate_history.get(mandate_id, [])
    return [
        MandateHistoryEntry(
            timestamp=h["timestamp"], status=MandateStatus(h["status"])
        )
        for h in history
    ]


@app.post("/import/properties", response_model=ImportResult)
def import_properties(
    file: UploadFile = File(...),
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    filename = file.filename or ""
    content = file.file.read()
    rows = []
    if filename.endswith(".csv"):
        try:
            text = content.decode("utf-8")
            reader = csv.DictReader(text.splitlines())
            rows = list(reader)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid CSV file")
    elif filename.endswith(".xls") or filename.endswith(".xlsx"):
        try:
            wb = load_workbook(BytesIO(content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid Excel file")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type")

    imported = 0
    errors: List[str] = []
    for idx, row in enumerate(rows, start=2):
        try:
            project_id = int(row.get("project_id"))
            project_name = row.get("project_name")
            stand_id = int(row.get("stand_id"))
            stand_name = row.get("stand_name") or ""
            size = float(row.get("size"))
            price = float(row.get("price"))
            if not project_name:
                raise ValueError("project_name is required")
        except (TypeError, ValueError) as e:
            errors.append(f"Row {idx}: {e}")
            continue

        project = repos.projects.get(project_id)
        if not project:
            repos.projects.add(Project(id=project_id, name=project_name))
        elif project.name != project_name:
            errors.append(
                f"Row {idx}: project name mismatch for ID {project_id}"
            )
            continue

        if repos.stands.get(stand_id):
            errors.append(f"Row {idx}: stand ID {stand_id} exists")
            continue

        stand = Stand(
            id=stand_id,
            project_id=project_id,
            name=stand_name,
            size=size,
            price=price,
        )
        repos.stands.add(stand)
        imported += 1

    return ImportResult(imported=imported, errors=errors)


# Reporting endpoints


@app.get("/reports/properties")
def properties_report(
    status: Optional[PropertyStatus] = None,
    format: str = "csv",
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    rows = generate_properties_report(repos, status)
    fieldnames = [
        "project_id",
        "project_name",
        "stand_id",
        "stand_name",
        "price",
        "status",
        "mandate_status",
    ]
    if format == "excel":
        data = list(rows)
        wb = Workbook()
        ws = wb.active
        ws.append(fieldnames)
        for row in data:
            ws.append([row[f] for f in fieldnames])
        stream = BytesIO()
        wb.save(stream)
        return Response(
            content=stream.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return StreamingResponse(stream_csv(rows, fieldnames), media_type="text/csv")


@app.get("/reports/mandates")
def mandates_report(
    status: Optional[MandateStatus] = None,
    format: str = "csv",
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    rows = generate_mandates_report(repos, status)
    fieldnames = [
        "project_id",
        "project_name",
        "stand_id",
        "stand_name",
        "agent",
        "status",
        "expiration_date",
    ]
    if format == "excel":
        data = list(rows)
        wb = Workbook()
        ws = wb.active
        ws.append(fieldnames)
        for row in data:
            ws.append([row[f] for f in fieldnames])
        stream = BytesIO()
        wb.save(stream)
        return Response(
            content=stream.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return StreamingResponse(stream_csv(rows, fieldnames), media_type="text/csv")


@app.get("/reports/loans")
def loans_report(
    status: Optional[SubmissionStatus] = None,
    format: str = "csv",
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    rows = generate_loans_report(repos, status)
    fieldnames = ["loan_id", "realtor", "account_id", "status", "decision"]
    if format == "excel":
        data = list(rows)
        wb = Workbook()
        ws = wb.active
        ws.append(fieldnames)
        for row in data:
            ws.append([row[f] for f in fieldnames])
        stream = BytesIO()
        wb.save(stream)
        return Response(
            content=stream.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return StreamingResponse(stream_csv(rows, fieldnames), media_type="text/csv")


# ---- Submission endpoints ----


def _ensure_owner(obj_realtor: str, agent: Agent):
    if agent.role != AgentRole.ADMIN and obj_realtor != agent.username:
        raise HTTPException(status_code=403, detail="Not authorized")


def _requirements_for_workflow(
    repos: Repositories, workflow: DocumentWorkflow
) -> List[DocumentRequirement]:
    requirements = [
        req for req in repos.document_requirements.list() if req.applies_to == workflow
    ]
    return sorted(requirements, key=lambda r: r.order)


def _validate_required_documents(
    workflow: DocumentWorkflow,
    provided: Dict[str, UploadedFile] | None,
    repos: Repositories,
) -> None:
    provided = provided or {}
    requirements = _requirements_for_workflow(repos, workflow)
    missing = [
        req.name
        for req in requirements
        if req.slug not in provided or not provided[req.slug].content
    ]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required documents: {', '.join(missing)}",
        )
    unexpected = sorted(
        key for key in provided.keys() if key not in {req.slug for req in requirements}
    )
    if unexpected:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown document fields: {', '.join(unexpected)}",
        )


def _allocate_requirement_id(repos: Repositories) -> int:
    next_id = repos.counters.get("next_document_requirement_id")
    if next_id is None:
        existing = repos.document_requirements.list()
        next_id = max((req.id for req in existing), default=0) + 1
    repos.counters.set("next_document_requirement_id", next_id + 1)
    return next_id


def _next_requirement_order(repos: Repositories, workflow: DocumentWorkflow) -> int:
    existing = _requirements_for_workflow(repos, workflow)
    if not existing:
        return 1
    return max(req.order for req in existing) + 1


def _slugify_requirement_name(name: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", name.strip().lower())
    normalized = normalized.strip("_") or "document"
    return normalized


def _derive_unique_slug(
    repos: Repositories,
    workflow: DocumentWorkflow,
    name: str,
    *,
    exclude_id: Optional[int] = None,
) -> str:
    base = _slugify_requirement_name(name)
    existing = {
        req.slug
        for req in repos.document_requirements.list()
        if req.applies_to == workflow and (exclude_id is None or req.id != exclude_id)
    }
    if base not in existing:
        return base
    suffix = 2
    while True:
        candidate = f"{base}_{suffix}"
        if candidate not in existing:
            return candidate
        suffix += 1


def _loan_ids_for_account(account_id: int, repos: Repositories) -> List[int]:
    return sorted(
        {
            loan.id
            for loan in repos.loan_applications.list()
            if loan.account_id == account_id
        }
    )


def _agreement_ids_for_loans(loan_ids: List[int], repos: Repositories) -> List[int]:
    loan_id_set = set(loan_ids)
    if not loan_id_set:
        return []
    return sorted(
        {
            agreement.id
            for agreement in repos.agreements.list()
            if agreement.loan_application_id in loan_id_set
        }
    )


def _ensure_agreement_for_application(
    application: LoanApplication, repos: Repositories
) -> None:
    if not application.property_id:
        return
    if repos.agreements.get(application.id):
        return
    stand = repos.stands.get(application.property_id)
    if not stand:
        return
    content = f"Agreement for property {stand.name} with loan {application.id}"
    timestamp = datetime.utcnow().isoformat()
    agreement = Agreement(
        id=application.id,
        loan_application_id=application.id,
        property_id=application.property_id,
        realtor=application.realtor,
        document=content,
        versions=[content],
        audit_log=[f"{timestamp}: generated"],
        status=AgreementStatus.DRAFT,
    )
    repos.agreements.add(agreement)


def _find_account_opening_by_number(
    account_number: str, repos: Repositories
) -> AccountOpening | None:
    for opening in repos.account_openings.list():
        if opening.account_number == account_number:
            return opening
    return None


def _sync_profile_from_account(
    account: AccountOpening, repos: Repositories
) -> CustomerProfile | None:
    if not account.account_number:
        return None
    loan_ids = _loan_ids_for_account(account.id, repos)
    agreement_ids = _agreement_ids_for_loans(loan_ids, repos)
    existing = repos.customer_profiles.get(account.account_number)
    now = datetime.utcnow()
    if existing:
        profile = existing.model_copy(
            update={
                "account_opening_id": account.id,
                "realtor": account.realtor,
                "loan_application_ids": loan_ids,
                "agreement_ids": agreement_ids,
                "account_number": account.account_number,
                "updated_at": now,
            }
        )
    else:
        profile = CustomerProfile(
            id=account.account_number,
            account_number=account.account_number,
            account_opening_id=account.id,
            realtor=account.realtor,
            loan_application_ids=loan_ids,
            agreement_ids=agreement_ids,
            last_inbound_email_at=None,
            deletion_requested=False,
            deletion_requested_at=None,
            deletion_requested_by=None,
            deletion_approved_at=None,
            deletion_approved_by=None,
            created_at=now,
            updated_at=now,
        )
    repos.customer_profiles.add(profile)
    return profile


def _touch_profile(
    profile: CustomerProfile, repos: Repositories, **updates: Any
) -> CustomerProfile:
    payload = dict(updates)
    payload["updated_at"] = datetime.utcnow()
    new_profile = profile.model_copy(update=payload)
    repos.customer_profiles.add(new_profile)
    return new_profile


def _normalize_document_key(filename: str, existing: set[str]) -> str:
    base = re.sub(r"[^a-z0-9]+", "-", filename.lower()).strip("-") or "attachment"
    candidate = base
    counter = 1
    while candidate in existing:
        candidate = f"{base}-{counter}"
        counter += 1
    return candidate


def _store_profile_documents(
    profile: CustomerProfile, attachments: List[UploadedFile], repos: Repositories
) -> CustomerProfile:
    if not attachments:
        return profile
    documents = dict(profile.documents)
    existing = set(documents.keys())
    for attachment in attachments:
        name = attachment.filename or "attachment"
        key = _normalize_document_key(name, existing)
        documents[key] = attachment
        existing.add(key)
    return _touch_profile(profile, repos, documents=documents)


def _get_profile_or_404(
    repos: Repositories, account_number: str
) -> CustomerProfile:
    profile = repos.customer_profiles.get(account_number)
    if not profile:
        raise HTTPException(status_code=404, detail="Customer profile not found")
    return profile


def _require_profile_access(profile: CustomerProfile, agent: Agent) -> None:
    if agent.role in (AgentRole.ADMIN, AgentRole.MANAGER):
        return
    if agent.role == AgentRole.AGENT and profile.realtor == agent.username:
        return
    raise HTTPException(status_code=403, detail="Not authorized to access profile")


async def _encode_upload_file(
    upload: UploadFile | StarletteUploadFile,
) -> UploadedFile:
    contents = await upload.read()
    encoded = base64.b64encode(contents).decode("utf-8")
    return UploadedFile(
        filename=upload.filename or "",
        content_type=upload.content_type or "application/octet-stream",
        content=encoded,
    )


async def _collect_required_documents(
    form: FormData,
    workflow: DocumentWorkflow,
    repos: Repositories,
) -> Dict[str, UploadedFile]:
    documents: Dict[str, UploadedFile] = {}
    requirements = _requirements_for_workflow(repos, workflow)
    for requirement in requirements:
        value = form.get(requirement.slug)
        if value is None:
            continue
        if not isinstance(value, (UploadFile, StarletteUploadFile)):
            raise HTTPException(
                status_code=400,
                detail=f"Invalid upload for requirement '{requirement.name}'",
            )
        documents[requirement.slug] = await _encode_upload_file(value)
    return documents


@app.post("/applications/offer", response_model=Offer)
async def upload_offer_application(
    request: Request,
    id: int = Form(...),
    realtor: str = Form(...),
    property_id: int = Form(...),
    details: Optional[str] = Form(None),
    file: UploadFile = File(...),
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if repos.offers.get(id):
        raise HTTPException(status_code=400, detail="Offer ID exists")
    if agent.username != realtor:
        raise HTTPException(status_code=403, detail="Cannot submit for another realtor")
    filename = file.filename or ""
    if not (filename.endswith(".pdf") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Unsupported file type")
    encoded = base64.b64encode(await file.read()).decode("utf-8")
    if not encoded:
        raise HTTPException(status_code=400, detail="Primary document cannot be empty")
    document = UploadedFile(
        filename=filename, content_type=file.content_type, content=encoded
    )
    form = await request.form()
    parsed_documents = await _collect_required_documents(
        form, DocumentWorkflow.OFFER, repos
    )
    _validate_required_documents(DocumentWorkflow.OFFER, parsed_documents, repos)
    offer = Offer(
        id=id,
        realtor=realtor,
        property_id=property_id,
        details=details,
        document=document,
        required_documents=parsed_documents,
    )
    repos.offers.add(offer)
    repos.notifications.append(f"Offer {id} submitted by {realtor}")
    _send_submission_email(
        subject=f"Offer submission #{id}",
        metadata={
            "Submission Type": "Offer",
            "Offer ID": id,
            "Realtor": realtor,
            "Property ID": property_id,
            "Details": details,
        },
        primary_document=offer.document,
        required_documents=offer.required_documents,
        repos=repos,
    )
    return offer


@app.post("/applications/property", response_model=PropertyApplication)
async def upload_property_application(
    request: Request,
    id: int = Form(...),
    realtor: str = Form(...),
    property_id: int = Form(...),
    details: Optional[str] = Form(None),
    file: UploadFile = File(...),
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if repos.applications.get(id):
        raise HTTPException(status_code=400, detail="Application ID exists")
    if agent.username != realtor:
        raise HTTPException(status_code=403, detail="Cannot submit for another realtor")
    filename = file.filename or ""
    if not (filename.endswith(".pdf") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Unsupported file type")
    encoded = base64.b64encode(await file.read()).decode("utf-8")
    if not encoded:
        raise HTTPException(status_code=400, detail="Primary document cannot be empty")
    document = UploadedFile(
        filename=filename, content_type=file.content_type, content=encoded
    )
    form = await request.form()
    parsed_documents = await _collect_required_documents(
        form, DocumentWorkflow.PROPERTY_APPLICATION, repos
    )
    _validate_required_documents(
        DocumentWorkflow.PROPERTY_APPLICATION, parsed_documents, repos
    )
    application = PropertyApplication(
        id=id,
        realtor=realtor,
        property_id=property_id,
        details=details,
        document=document,
        required_documents=parsed_documents,
    )
    repos.applications.add(application)
    repos.notifications.append(
        f"Property application {id} submitted by {realtor}"
    )
    return application


@app.post("/applications/account", response_model=AccountOpening)
async def upload_account_opening(
    request: Request,
    id: int = Form(...),
    realtor: str = Form(...),
    details: Optional[str] = Form(None),
    file: UploadFile = File(...),
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if repos.account_openings.get(id):
        raise HTTPException(status_code=400, detail="Request ID exists")
    if agent.username != realtor:
        raise HTTPException(status_code=403, detail="Cannot submit for another realtor")
    filename = file.filename or ""
    if not (filename.endswith(".pdf") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Unsupported file type")
    encoded = base64.b64encode(await file.read()).decode("utf-8")
    if not encoded:
        raise HTTPException(status_code=400, detail="Primary document cannot be empty")
    document = UploadedFile(
        filename=filename, content_type=file.content_type, content=encoded
    )
    form = await request.form()
    parsed_documents = await _collect_required_documents(
        form, DocumentWorkflow.ACCOUNT_OPENING, repos
    )
    _validate_required_documents(
        DocumentWorkflow.ACCOUNT_OPENING, parsed_documents, repos
    )
    request = AccountOpening(
        id=id,
        realtor=realtor,
        details=details,
        document=document,
        required_documents=parsed_documents,
        status=SubmissionStatus.SUBMITTED,
    )
    repos.account_openings.add(request)
    repos.notifications.append(
        f"Account opening {id} submitted by {realtor}"
    )
    return request


@app.post("/offers", response_model=Offer)
def submit_offer(
    offer: Offer,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if repos.offers.get(offer.id):
        raise HTTPException(status_code=400, detail="Offer ID exists")
    if agent.username != offer.realtor:
        raise HTTPException(status_code=403, detail="Cannot submit for another realtor")
    _validate_required_documents(
        DocumentWorkflow.OFFER, offer.required_documents, repos
    )
    sanitized_offer = offer.model_copy(
        update={"status": SubmissionStatus.SUBMITTED}
    )
    repos.offers.add(sanitized_offer)
    repos.notifications.append(
        f"Offer {sanitized_offer.id} submitted by {sanitized_offer.realtor}"
    )
    _send_submission_email(
        subject=f"Offer submission #{sanitized_offer.id}",
        metadata={
            "Submission Type": "Offer",
            "Offer ID": sanitized_offer.id,
            "Realtor": sanitized_offer.realtor,
            "Property ID": sanitized_offer.property_id,
            "Details": sanitized_offer.details,
        },
        primary_document=sanitized_offer.document,
        required_documents=sanitized_offer.required_documents,
        repos=repos,
    )
    return sanitized_offer


@app.get("/offers/{offer_id}", response_model=Offer)
def get_offer(
    offer_id: int,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    offer = repos.offers.get(offer_id)
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    _ensure_owner(offer.realtor, agent)
    return offer


@app.put("/offers/{offer_id}/status", response_model=Offer)
def update_offer_status(
    offer_id: int,
    update: StatusUpdate,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    offer = repos.offers.get(offer_id)
    if not offer:
        raise HTTPException(status_code=404, detail="Offer not found")
    offer.status = update.status
    repos.offers.add(offer)
    return offer


@app.post("/property-applications", response_model=PropertyApplication)
def submit_property_application(
    application: PropertyApplication,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if repos.applications.get(application.id):
        raise HTTPException(status_code=400, detail="Application ID exists")
    if agent.username != application.realtor:
        raise HTTPException(status_code=403, detail="Cannot submit for another realtor")
    _validate_required_documents(
        DocumentWorkflow.PROPERTY_APPLICATION,
        application.required_documents,
        repos,
    )
    sanitized_application = application.model_copy(
        update={"status": SubmissionStatus.SUBMITTED}
    )
    repos.applications.add(sanitized_application)
    repos.notifications.append(
        f"Property application {sanitized_application.id} submitted by {sanitized_application.realtor}"
    )
    return sanitized_application


@app.get("/property-applications", response_model=List[PropertyApplication])
def list_property_applications(
    status: Optional[SubmissionStatus] = None,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    applications = repos.applications.list()
    if status:
        applications = [app for app in applications if app.status == status]
    return applications


@app.post("/property-applications/{app_id}/approve", response_model=PropertyApplication)
def approve_property_application(
    app_id: int,
    agent: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    application = repos.applications.get(app_id)
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    if application.status != SubmissionStatus.SUBMITTED:
        raise HTTPException(status_code=400, detail="Only submitted applications can be approved")
    application.status = SubmissionStatus.MANAGER_APPROVED
    repos.applications.add(application)
    repos.notifications.append(
        f"Property application {app_id} approved by {agent.username}"
    )
    _send_submission_email(
        subject=f"Property application #{application.id}",
        metadata={
            "Submission Type": "Property Application",
            "Application ID": application.id,
            "Realtor": application.realtor,
            "Property ID": application.property_id,
            "Details": application.details,
        },
        primary_document=application.document,
        required_documents=application.required_documents,
        repos=repos,
    )
    return application


@app.get("/property-applications/{app_id}", response_model=PropertyApplication)
def get_property_application(
    app_id: int,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    application = repos.applications.get(app_id)
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    _ensure_owner(application.realtor, agent)
    return application


@app.put("/property-applications/{app_id}/status", response_model=PropertyApplication)
def update_property_application_status(
    app_id: int,
    update: StatusUpdate,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    application = repos.applications.get(app_id)
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    application.status = update.status
    repos.applications.add(application)
    return application


@app.post("/account-openings", response_model=AccountOpening)
def submit_account_opening(
    request: AccountOpening,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if repos.account_openings.get(request.id):
        raise HTTPException(status_code=400, detail="Request ID exists")
    if agent.username != request.realtor:
        raise HTTPException(status_code=403, detail="Cannot submit for another realtor")
    _validate_required_documents(
        DocumentWorkflow.ACCOUNT_OPENING, request.required_documents, repos
    )
    sanitized_request = request.model_copy(
        update={
            "status": SubmissionStatus.SUBMITTED,
            "account_number": None,
            "deposit_threshold": None,
            "deposits": [],
        }
    )
    repos.account_openings.add(sanitized_request)
    repos.notifications.append(
        f"Account opening {sanitized_request.id} submitted by {sanitized_request.realtor}"
    )
    return sanitized_request


@app.post("/account-openings/{req_id}/approve", response_model=AccountOpening)
def approve_account_opening(
    req_id: int,
    agent: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    request = repos.account_openings.get(req_id)
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    if request.status != SubmissionStatus.SUBMITTED:
        raise HTTPException(status_code=400, detail="Only submitted requests can be approved")
    request.status = SubmissionStatus.MANAGER_APPROVED
    repos.account_openings.add(request)
    repos.notifications.append(
        f"Account opening {req_id} approved by {agent.username}"
    )
    _send_submission_email(
        subject=f"Account opening #{request.id}",
        metadata={
            "Submission Type": "Account Opening",
            "Request ID": request.id,
            "Realtor": request.realtor,
            "Details": request.details,
        },
        primary_document=request.document,
        required_documents=request.required_documents,
        repos=repos,
    )
    return request


@app.post("/applications/account/{req_id}/approve", response_model=AccountOpening)
def approve_uploaded_account_opening(
    req_id: int,
    agent: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    return approve_account_opening(req_id=req_id, agent=agent, repos=repos)


@app.get("/account-openings", response_model=List[AccountOpening])
def list_account_openings(
    status: Optional[SubmissionStatus] = None,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    openings = repos.account_openings.list()
    if status:
        openings = [o for o in openings if o.status == status]
    return openings


@app.get("/account-openings/{req_id}", response_model=AccountOpening)
def get_account_opening(
    req_id: int,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    request = repos.account_openings.get(req_id)
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    if agent.role not in (AgentRole.ADMIN, AgentRole.MANAGER) and request.realtor != agent.username:
        raise HTTPException(status_code=403, detail="Not authorized")
    return request


@app.put("/account-openings/{req_id}/status", response_model=AccountOpening)
def update_account_opening_status(
    req_id: int,
    update: StatusUpdate,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    request = repos.account_openings.get(req_id)
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    request.status = update.status
    repos.account_openings.add(request)
    return request


@app.put("/account-openings/{req_id}/open", response_model=AccountOpening)
def open_account(
    req_id: int,
    details: AccountSetup,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    request = repos.account_openings.get(req_id)
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    if request.status not in (
        SubmissionStatus.MANAGER_APPROVED,
        SubmissionStatus.IN_PROGRESS,
        SubmissionStatus.COMPLETED,
    ):
        raise HTTPException(
            status_code=400,
            detail="Account opening must be manager approved before setup",
        )
    if details.deposit_threshold <= 0:
        raise HTTPException(status_code=400, detail="Deposit threshold must be positive")
    request.account_number = details.account_number
    request.deposit_threshold = details.deposit_threshold
    request.status = SubmissionStatus.IN_PROGRESS
    repos.account_openings.add(request)
    _sync_profile_from_account(request, repos)
    return request


@app.post("/account-openings/{req_id}/deposit", response_model=AccountOpening)
def record_deposit(
    req_id: int,
    deposit: Deposit,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    request = repos.account_openings.get(req_id)
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    if deposit.amount <= 0:
        raise HTTPException(status_code=400, detail="Deposit amount must be positive")
    request.deposits.append(deposit.amount)
    if request.deposit_threshold is not None and sum(request.deposits) >= request.deposit_threshold:
        request.status = SubmissionStatus.COMPLETED
    repos.account_openings.add(request)
    return request


@app.get("/accounts/deposits/pending", response_model=List[AccountOpening])
def list_pending_deposits(
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    openings = repos.account_openings.list()
    return [
        o
        for o in openings
        if o.status
        in (
            SubmissionStatus.SUBMITTED,
            SubmissionStatus.MANAGER_APPROVED,
            SubmissionStatus.IN_PROGRESS,
        )
    ]


@app.get("/accounts/deposits/imported", response_model=List[ImportedDepositAccount])
def list_imported_deposits(
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    return repos.imported_deposit_accounts.list()


@app.get("/accounts/deposits/imported/{record_id}", response_model=ImportedDepositAccount)
def get_imported_deposit(
    record_id: str,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    record = repos.imported_deposit_accounts.get(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Imported deposit account not found")
    return record


@app.post("/accounts/deposits/{req_id}/open", response_model=AccountOpening)
def open_deposit_account(
    req_id: int,
    details: AccountSetup,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    request = repos.account_openings.get(req_id)
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    if request.status not in (
        SubmissionStatus.MANAGER_APPROVED,
        SubmissionStatus.IN_PROGRESS,
        SubmissionStatus.COMPLETED,
    ):
        raise HTTPException(
            status_code=400,
            detail="Account opening must be manager approved before setup",
        )
    if details.deposit_threshold <= 0:
        raise HTTPException(status_code=400, detail="Deposit threshold must be positive")
    request.account_number = details.account_number
    request.deposit_threshold = details.deposit_threshold
    request.status = SubmissionStatus.IN_PROGRESS
    repos.account_openings.add(request)
    return request


@app.post("/accounts/deposits/{req_id}/deposit", response_model=AccountOpening)
def record_account_deposit(
    req_id: int,
    deposit: Deposit,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    request = repos.account_openings.get(req_id)
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    if deposit.amount <= 0:
        raise HTTPException(status_code=400, detail="Deposit amount must be positive")
    request.deposits.append(deposit.amount)
    if request.deposit_threshold is not None and sum(request.deposits) >= request.deposit_threshold:
        request.status = SubmissionStatus.COMPLETED
    repos.account_openings.add(request)
    return request


# ---- Customer profile endpoints ----


@app.get("/profiles/{account_number}", response_model=CustomerProfile)
def get_customer_profile(
    account_number: str,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    profile = _get_profile_or_404(repos, account_number)
    account = _find_account_opening_by_number(account_number, repos)
    if account:
        refreshed = _sync_profile_from_account(account, repos)
        if refreshed:
            profile = refreshed
    _require_profile_access(profile, agent)
    return profile


@app.post("/profiles/{account_number}/request-deletion", response_model=CustomerProfile)
def request_customer_profile_deletion(
    account_number: str,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    profile = _get_profile_or_404(repos, account_number)
    _require_profile_access(profile, agent)
    if agent.role != AgentRole.AGENT:
        raise HTTPException(status_code=403, detail="Only agents can request deletion")
    updated = _touch_profile(
        profile,
        repos,
        deletion_requested=True,
        deletion_requested_at=datetime.utcnow(),
        deletion_requested_by=agent.username,
        deletion_approved_at=None,
        deletion_approved_by=None,
    )
    repos.notifications.append(
        f"Deletion requested for profile {account_number} by {agent.username}"
    )
    return updated


@app.delete("/profiles/{account_number}", status_code=204)
def delete_customer_profile(
    account_number: str,
    agent: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    profile = repos.customer_profiles.get(account_number)
    if not profile:
        raise HTTPException(status_code=404, detail="Customer profile not found")
    if not profile.deletion_requested:
        raise HTTPException(status_code=400, detail="No pending deletion request")
    repos.customer_profiles.delete(account_number)
    repos.notifications.append(
        f"Customer profile {account_number} deleted by {agent.username}"
    )
    return Response(status_code=204)


@app.post("/loan-applications", response_model=LoanApplication)
def submit_loan_application(
    application: LoanApplication,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if repos.loan_applications.get(application.id):
        raise HTTPException(status_code=400, detail="Loan application ID exists")
    if agent.username != application.realtor:
        raise HTTPException(status_code=403, detail="Cannot submit for another realtor")
    account = repos.account_openings.get(application.account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    if account.realtor != agent.username or account.realtor != application.realtor:
        raise HTTPException(status_code=403, detail="Cannot submit for another realtor")
    if account.status != SubmissionStatus.COMPLETED:
        raise HTTPException(
            status_code=400, detail="Deposits not sufficient for loan application"
        )
    _validate_required_documents(
        DocumentWorkflow.LOAN_APPLICATION,
        application.required_documents,
        repos,
    )
    if application.property_id and not repos.stands.get(application.property_id):
        raise HTTPException(status_code=404, detail="Property not found")
    sanitized_application = application.model_copy(
        update={
            "status": SubmissionStatus.SUBMITTED,
            "decision": None,
            "reason": None,
            "loan_account_number": None,
        }
    )
    repos.loan_applications.add(sanitized_application)
    _sync_profile_from_account(account, repos)
    repos.notifications.append(
        f"Loan application {sanitized_application.id} submitted by {sanitized_application.realtor}"
    )
    _send_submission_email(
        subject=f"Loan application #{sanitized_application.id}",
        metadata={
            "Submission Type": "Loan Application",
            "Application ID": sanitized_application.id,
            "Realtor": sanitized_application.realtor,
            "Account ID": sanitized_application.account_id,
            "Property ID": sanitized_application.property_id,
        },
        primary_document=None,
        required_documents=sanitized_application.required_documents,
        repos=repos,
    )
    return sanitized_application


@app.post("/automation/loan-team-reply", response_model=LoanApplication)
def process_loan_team_reply(
    reply: LoanTeamReply,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    application = repos.loan_applications.get(reply.loan_application_id)
    if not application:
        raise HTTPException(status_code=404, detail="Loan application not found")
    decision = _parse_loan_team_decision(reply.body)
    if not decision:
        raise HTTPException(status_code=400, detail="Unable to determine decision from message")
    application.decision = decision
    if decision == LoanDecision.APPROVED:
        application.status = SubmissionStatus.COMPLETED
        application.reason = None
        _ensure_agreement_for_application(application, repos)
    else:
        application.status = SubmissionStatus.REJECTED
        application.reason = reply.body.strip() or "Declined by loan team"
    repos.loan_applications.add(application)
    repos.notifications.append(
        f"Loan application {application.id} updated from loan team reply"
    )
    account = repos.account_openings.get(application.account_id)
    if account and account.account_number:
        profile = repos.customer_profiles.get(account.account_number)
        if not profile:
            profile = _sync_profile_from_account(account, repos)
        if profile:
            _store_profile_documents(profile, reply.attachments, repos)
    return application


@app.get("/loan-applications/{app_id}", response_model=LoanApplication)
def get_loan_application(
    app_id: int,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    application = repos.loan_applications.get(app_id)
    if not application:
        raise HTTPException(status_code=404, detail="Loan application not found")
    _ensure_owner(application.realtor, agent)
    return application


@app.get("/loan-applications", response_model=List[LoanApplication])
def list_loan_applications(
    status: Optional[SubmissionStatus] = None,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    apps = repos.loan_applications.list()
    if status:
        apps = [a for a in apps if a.status == status]
    return apps


@app.put("/loan-applications/{app_id}/decision", response_model=LoanApplication)
def decide_loan_application(
    app_id: int,
    decision: LoanDecisionUpdate,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    application = repos.loan_applications.get(app_id)
    if not application:
        raise HTTPException(status_code=404, detail="Loan application not found")
    application.decision = decision.decision
    application.reason = decision.reason
    if decision.decision == LoanDecision.APPROVED:
        application.status = SubmissionStatus.COMPLETED
        if application.property_id and repos.agreements.get(application.id):
            raise HTTPException(status_code=400, detail="Agreement ID exists")
        _ensure_agreement_for_application(application, repos)
    else:
        application.status = SubmissionStatus.REJECTED
    repos.loan_applications.add(application)
    return application


@app.post("/loans", response_model=Loan)
def submit_loan(
    loan: Loan,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if repos.loans.get(loan.id):
        raise HTTPException(status_code=400, detail="Loan ID exists")
    if agent.username != loan.borrower:
        raise HTTPException(status_code=403, detail="Cannot submit for another borrower")
    sanitized_loan = loan.model_copy(update={"status": LoanStatus.SUBMITTED, "reason": None})
    repos.loans.add(sanitized_loan)
    return sanitized_loan


@app.get("/loans/pending", response_model=List[Loan])
def list_pending_loans(
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    loans = [
        l
        for l in repos.loans.list()
        if l.status in (LoanStatus.SUBMITTED, LoanStatus.UNDER_REVIEW)
    ]
    result = []
    for loan in loans:
        if loan.status == LoanStatus.SUBMITTED:
            loan.status = LoanStatus.UNDER_REVIEW
            repos.loans.add(loan)
        result.append(loan)
    return result


@app.post("/loans/{loan_id}/decision", response_model=Loan)
def decide_loan(
    loan_id: int,
    decision: LoanDecisionRequest,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    loan = repos.loans.get(loan_id)
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    if decision.decision not in (LoanStatus.APPROVED, LoanStatus.REJECTED):
        raise HTTPException(status_code=400, detail="Invalid decision")
    loan.status = decision.decision
    loan.reason = decision.reason
    repos.loans.add(loan)
    return loan


@app.get("/notifications", response_model=List[str])
def list_notifications(
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    return repos.notifications.list()


class LoanAccountRequest(BaseModel):
    agreement_id: int


@app.get("/loan-accounts/imported", response_model=List[ImportedLoanAccount])
def list_imported_loan_accounts(
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    return repos.imported_loan_accounts.list()


@app.get("/loan-accounts/imported/{record_id}", response_model=ImportedLoanAccount)
def get_imported_loan_account(
    record_id: str,
    _: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    record = repos.imported_loan_accounts.get(record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Imported loan account not found")
    return record


@app.post("/loan-accounts", response_model=dict)
def create_loan_account(
    payload: LoanAccountRequest,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    agreement = repos.agreements.get(payload.agreement_id)
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    if agreement.status != AgreementStatus.SIGNED:
        raise HTTPException(status_code=400, detail="Agreement not fully signed")
    loan_application = repos.loan_applications.get(agreement.loan_application_id)
    if not loan_application:
        raise HTTPException(status_code=404, detail="Loan application not found")
    if loan_application.loan_account_number:
        raise HTTPException(status_code=400, detail="Loan account already created")

    account_number = f"LN{uuid.uuid4().hex[:8].upper()}"
    loan_application.loan_account_number = account_number
    repos.loan_applications.add(loan_application)

    accounts = repos.customer_loan_accounts.get(loan_application.realtor, [])
    accounts.append(account_number)
    repos.customer_loan_accounts.set(loan_application.realtor, accounts)

    stand = repos.stands.get(agreement.property_id)
    if not stand:
        raise HTTPException(status_code=404, detail="Property not found")
    if stand.status == PropertyStatus.SOLD:
        raise HTTPException(status_code=400, detail="Property already sold")
    stand.status = PropertyStatus.SOLD
    repos.stands.add(stand)

    return {"account_number": account_number}


@app.get("/loan-accounts/{realtor}", response_model=List[str])
def list_loan_accounts(
    realtor: str,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    if agent.role not in (AgentRole.ADMIN, AgentRole.MANAGER) and agent.username != realtor:
        raise HTTPException(status_code=403, detail="Not authorized")
    return repos.customer_loan_accounts.get(realtor, [])


@app.get("/dashboard")
def dashboard(
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    data = {}
    if agent.role in (AgentRole.MANAGER, AgentRole.ADMIN):
        stands_list = repos.stands.list()
        prop_counts = {
            status.value: sum(1 for s in stands_list if s.status == status)
            for status in PropertyStatus
        }
        mandate_counts = {
            status.value: sum(
                1
                for s in stands_list
                if s.mandate and s.mandate.status == status
            )
            for status in MandateStatus
        }
        data["property_status"] = prop_counts
        data["mandates"] = mandate_counts
    if agent.role in (AgentRole.COMPLIANCE, AgentRole.ADMIN):
        openings = repos.account_openings.list()
        total_deposits = sum(sum(req.deposits) for req in openings)
        apps = repos.loan_applications.list()
        approvals = sum(
            1 for app in apps if app.decision == LoanDecision.APPROVED
        )
        rejections = sum(
            1 for app in apps if app.decision == LoanDecision.REJECTED
        )
        data["deposits"] = total_deposits
        data["loan_approvals"] = {"approved": approvals, "rejected": rejections}
    return data


@app.get("/audit-log", response_model=List[AuditEntry])
def get_audit_log(
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    user: Optional[str] = None,
    action: Optional[str] = None,
    _: Agent = Depends(require_compliance),
    repos: Repositories = Depends(get_repositories),
):
    entries = []
    for entry in repos.audit_log.list():
        parts = entry.split(" - ")
        if len(parts) != 4:
            continue
        ts_str, username, act, status_str = parts
        ts = datetime.fromisoformat(ts_str)
        if start and ts < start:
            continue
        if end and ts > end:
            continue
        if user and username != user:
            continue
        if action and action not in act:
            continue
        entries.append(
            AuditEntry(
                timestamp=ts_str, user=username, action=act, status=int(status_str)
            )
        )
    return entries


# ---- Agreement endpoints ----


@app.get("/agreements", response_model=List[Agreement])
def list_agreements(
    status: Optional[AgreementStatus] = None,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    agreements = repos.agreements.list()
    if agent.role == AgentRole.AGENT:
        agreements = [a for a in agreements if a.realtor == agent.username]
    if status:
        agreements = [a for a in agreements if a.status == status]
    return agreements


@app.post("/agreements", response_model=Agreement)
def create_agreement(
    data: AgreementCreate,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    if repos.agreements.get(data.id):
        raise HTTPException(status_code=400, detail="Agreement ID exists")
    loan = repos.loan_applications.get(data.loan_application_id)
    if not loan:
        raise HTTPException(status_code=404, detail="Loan application not found")
    account = repos.account_openings.get(loan.account_id)
    if not account:
        raise HTTPException(status_code=404, detail="Account opening not found")
    stand = repos.stands.get(data.property_id)
    if not stand:
        raise HTTPException(status_code=404, detail="Property not found")
    content = f"Agreement for property {stand.name} with loan {loan.id}"
    timestamp = datetime.utcnow().isoformat()
    agreement = Agreement(
        id=data.id,
        loan_application_id=data.loan_application_id,
        property_id=data.property_id,
        realtor=loan.realtor,
        document=content,
        versions=[content],
        audit_log=[f"{timestamp}: generated"],
        status=AgreementStatus.DRAFT,
    )
    repos.agreements.add(agreement)
    _sync_profile_from_account(account, repos)
    return agreement


@app.get("/agreements/{agreement_id}", response_model=Agreement)
def get_agreement(
    agreement_id: int,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    agreement = repos.agreements.get(agreement_id)
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    _ensure_owner(agreement.realtor, agent)
    return agreement


@app.get("/agreements/{agreement_id}/download")
def download_agreement(
    agreement_id: int,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    agreement = repos.agreements.get(agreement_id)
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    _ensure_owner(agreement.realtor, agent)
    if agreement.document_file:
        data = base64.b64decode(agreement.document_file.content)
        media_type = (
            agreement.document_file.content_type or "application/octet-stream"
        )
        extension = media_type.split("/")[-1] if "/" in media_type else "bin"
        filename = (
            agreement.document_file.filename
            or f"agreement-{agreement_id}.{extension}"
        )
    else:
        data = agreement.document.encode("utf-8")
        filename = f"agreement-{agreement_id}.txt"
        media_type = "text/plain"
    headers = {
        "Content-Disposition": f"attachment; filename=\"{filename}\""
    }
    return StreamingResponse(BytesIO(data), media_type=media_type, headers=headers)


@app.get("/agreements/{agreement_id}/document")
def view_agreement_document(
    agreement_id: int,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    agreement = repos.agreements.get(agreement_id)
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    _ensure_owner(agreement.realtor, agent)
    if agreement.document_file:
        data = base64.b64decode(agreement.document_file.content)
        return Response(content=data, media_type=agreement.document_file.content_type)
    return Response(content=agreement.document, media_type="text/plain")


@app.post(
    "/agreements/{agreement_id}/customer-upload-file", response_model=Agreement
)
async def upload_customer_agreement_file(
    agreement_id: int,
    file: UploadFile = File(...),
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    agreement = repos.agreements.get(agreement_id)
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    _ensure_owner(agreement.realtor, agent)
    encoded = await _encode_upload_file(file)
    agreement.version_files.append(encoded)
    agreement.document_file = encoded
    agreement.status = AgreementStatus.PARTIALLY_SIGNED
    timestamp = datetime.utcnow().isoformat()
    agreement.audit_log.append(
        f"{timestamp}: customer upload by {agent.username}"
    )
    repos.agreements.add(agreement)
    loan = repos.loan_applications.get(agreement.loan_application_id)
    if loan:
        account = repos.account_openings.get(loan.account_id)
        if account and account.account_number:
            profile = repos.customer_profiles.get(account.account_number)
            if not profile:
                profile = _sync_profile_from_account(account, repos)
            if profile:
                _store_profile_documents(profile, [encoded], repos)
    return agreement


@app.post(
    "/agreements/{agreement_id}/manager-upload-file", response_model=Agreement
)
async def upload_manager_agreement_file(
    agreement_id: int,
    file: UploadFile = File(...),
    agent: Agent = Depends(require_management),
    repos: Repositories = Depends(get_repositories),
):
    agreement = repos.agreements.get(agreement_id)
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    encoded = await _encode_upload_file(file)
    agreement.version_files.append(encoded)
    agreement.document_file = encoded
    agreement.status = AgreementStatus.SIGNED
    timestamp = datetime.utcnow().isoformat()
    agreement.audit_log.append(
        f"{timestamp}: manager upload by {agent.username}"
    )
    repos.agreements.add(agreement)
    loan = repos.loan_applications.get(agreement.loan_application_id)
    if loan:
        account = repos.account_openings.get(loan.account_id)
        if account and account.account_number:
            profile = repos.customer_profiles.get(account.account_number)
            if not profile:
                profile = _sync_profile_from_account(account, repos)
            if profile:
                _store_profile_documents(profile, [encoded], repos)
    return agreement


@app.post("/agreements/{agreement_id}/sign", response_model=Agreement)
def sign_agreement(
    agreement_id: int,
    payload: AgreementSign,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    agreement = repos.agreements.get(agreement_id)
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    timestamp = datetime.utcnow().isoformat()
    if agent.role == AgentRole.ADMIN:
        agreement.bank_document_url = payload.document_url
        agreement.audit_log.append(
            f"{timestamp}: bank signed by {agent.username}"
        )
    else:
        if agent.username != agreement.realtor:
            raise HTTPException(status_code=403, detail="Agent not authorized to sign")
        agreement.customer_document_url = payload.document_url
        agreement.audit_log.append(
            f"{timestamp}: customer signed by {agent.username}"
        )

    if agreement.bank_document_url and agreement.customer_document_url:
        agreement.status = AgreementStatus.SIGNED
        message = (
            f"Agreement {agreement_id} fully signed; notify Loan Accounts Opening Team"
        )
        if message not in repos.notifications.list():
            repos.notifications.append(message)
    else:
        agreement.status = AgreementStatus.PARTIALLY_SIGNED

    repos.agreements.add(agreement)

    return agreement


@app.post("/agreements/{agreement_id}/upload", response_model=Agreement)
def upload_agreement(
    agreement_id: int,
    upload: AgreementUpload,
    agent: Agent = Depends(get_current_agent),
    repos: Repositories = Depends(get_repositories),
):
    agreement = repos.agreements.get(agreement_id)
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    _ensure_owner(agreement.realtor, agent)
    agreement.versions.append(upload.document)
    agreement.document = upload.document
    timestamp = datetime.utcnow().isoformat()
    role = "bank" if agent.role == AgentRole.ADMIN else "customer"
    agreement.audit_log.append(f"{timestamp}: {role} uploaded new version")
    repos.agreements.add(agreement)
    return agreement


@app.put("/agreements/{agreement_id}/execute", response_model=Agreement)
def execute_agreement(
    agreement_id: int,
    execution: AgreementExecution,
    _: Agent = Depends(require_admin),
    repos: Repositories = Depends(get_repositories),
):
    agreement = repos.agreements.get(agreement_id)
    if not agreement:
        raise HTTPException(status_code=404, detail="Agreement not found")
    if agreement.status != AgreementStatus.SIGNED:
        raise HTTPException(status_code=400, detail="Agreement not fully signed")
    loan_application = repos.loan_applications.get(agreement.loan_application_id)
    loan_application.loan_account_number = execution.loan_account_number
    repos.loan_applications.add(loan_application)
    realtor = loan_application.realtor
    accounts = repos.customer_loan_accounts.get(realtor, [])
    accounts.append(execution.loan_account_number)
    repos.customer_loan_accounts.set(realtor, accounts)
    stand = repos.stands.get(agreement.property_id)
    stand.status = PropertyStatus.SOLD
    repos.stands.add(stand)
    repos.agreements.add(agreement)
    return agreement
